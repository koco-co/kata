from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from data_assets_playwright_web_ui.domains.data_quality.sql_merge_optimization.assertions import (
    SqlMergeAssertionError,
    assert_download_matches_visible_rows,
    assert_shared_dirty_table,
    assert_sql_topology,
    read_xlsx_snapshot,
)
from data_assets_playwright_web_ui.domains.data_quality.sql_merge_optimization.model import (
    MergeMode,
    SqlRuleIdentity,
    SqlTopologyExpectation,
    TableSnapshot,
)

if TYPE_CHECKING:
    from pathlib import Path

_PARTIAL_SCAN_GROUP_COUNT = 3


def _rule_identities(count: int) -> tuple[SqlRuleIdentity, ...]:
    return tuple(
        SqlRuleIdentity(index=index, token=str(1000 + index)) for index in range(1, count + 1)
    )


def test_shared_dirty_table_accepts_multiple_statements_with_one_target() -> None:
    sql = """
    INSERT INTO schema.dirty_shared SELECT 1;
    INSERT OVERWRITE TABLE `schema`.`dirty_shared` SELECT 2;
    """

    assert assert_shared_dirty_table(sql) == "schema.dirty_shared"


@pytest.mark.parametrize(
    "sql",
    [
        "SELECT 1",
        "INSERT INTO schema.dirty_a SELECT 1; INSERT INTO schema.dirty_b SELECT 2",
    ],
)
def test_shared_dirty_table_rejects_missing_or_split_targets(sql: str) -> None:
    with pytest.raises(SqlMergeAssertionError):
        assert_shared_dirty_table(sql)


def test_xlsx_snapshot_is_bounded_and_matches_visible_business_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["id", "name"])
    sheet.append([1, "alpha"])
    sheet.append([2, "beta"])
    download = tmp_path / "details.xlsx"
    workbook.save(download)

    downloaded = read_xlsx_snapshot(download, maximum_rows=100)
    visible = TableSnapshot(headers=("id", "name"), rows=(("1", "alpha"), ("2", "beta")))

    assert downloaded == visible
    assert_download_matches_visible_rows(downloaded, visible)


def test_xlsx_snapshot_rejects_more_than_the_canonical_100_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["id"])
    for value in range(101):
        sheet.append([value])
    download = tmp_path / "too-many.xlsx"
    workbook.save(download)

    with pytest.raises(SqlMergeAssertionError, match="at most 100"):
        read_xlsx_snapshot(download, maximum_rows=100)


def test_xlsx_snapshot_rejects_oversized_download_before_reading(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    oversized = tmp_path / "oversized.xlsx"
    with oversized.open("wb") as stream:
        stream.truncate(10 * 1024 * 1024 + 1)
    path_type = type(oversized)

    def fail_if_opened(*_args: object, **_kwargs: object) -> None:
        pytest.fail("oversized workbook payload must not be opened")

    monkeypatch.setattr(path_type, "open", fail_if_opened)

    with pytest.raises(SqlMergeAssertionError, match="fixed size limit"):
        read_xlsx_snapshot(oversized, maximum_rows=100)


def test_xlsx_snapshot_rejects_malformed_zip_without_disclosing_path(tmp_path: Path) -> None:
    malformed = tmp_path / "sensitive-customer-path.xlsx"
    malformed.write_bytes(b"not-an-xlsx-zip")

    with pytest.raises(SqlMergeAssertionError, match="readable XLSX") as caught:
        read_xlsx_snapshot(malformed, maximum_rows=100)

    assert str(malformed) not in str(caught.value)


def test_xlsx_snapshot_rejects_high_ratio_zip_bomb_without_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    archive = BytesIO()
    with ZipFile(archive, "w", compression=ZIP_DEFLATED) as zipped:
        zipped.writestr("xl/worksheets/sheet1.xml", b"A" * (2 * 1024 * 1024))
    download = tmp_path / "compressed-bomb.xlsx"
    download.write_bytes(archive.getvalue())

    def fail_if_loaded(*_args: object, **_kwargs: object) -> None:
        pytest.fail("unsafe archive must be rejected before openpyxl")

    monkeypatch.setattr(
        "data_assets_playwright_web_ui.domains.data_quality.sql_merge_optimization."
        "assertions.load_workbook",
        fail_if_loaded,
    )

    with pytest.raises(SqlMergeAssertionError, match="archive safety limits") as caught:
        read_xlsx_snapshot(download, maximum_rows=100)

    assert str(download) not in str(caught.value)


def test_xlsx_snapshot_stops_streaming_at_first_excess_data_row(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["id"])
    download = tmp_path / "bounded-stream.xlsx"
    workbook.save(download)

    class FakeSheet:
        def iter_rows(self, *, values_only: bool) -> object:
            assert values_only is True
            yield ("id",)
            for value in range(101):
                yield (value,)
            pytest.fail("row iterator consumed data after the first excess row")

    class FakeWorkbook:
        active = FakeSheet()

        @staticmethod
        def close() -> None:
            return None

    def fake_load_workbook(*_args: object, **_kwargs: object) -> FakeWorkbook:
        return FakeWorkbook()

    monkeypatch.setattr(
        "data_assets_playwright_web_ui.domains.data_quality.sql_merge_optimization."
        "assertions.load_workbook",
        fake_load_workbook,
    )

    with pytest.raises(SqlMergeAssertionError, match="at most 100"):
        read_xlsx_snapshot(download, maximum_rows=100)


def test_sql_topology_accepts_one_sampled_merged_scan_and_shared_dirty_target() -> None:
    sql = """
    CREATE TABLE schema.test_table_15862_c0019_temp_sample_table AS
      SELECT * FROM schema.test_table_15862_c0019 WHERE rand() <= 0.5 AND dt='2026-08-04';
    INSERT INTO schema.dirty_shared
      SELECT SUM(CASE WHEN id > 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002,
             SUM(CASE WHEN address = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1003,
             SUM(CASE WHEN age < 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1004
      FROM schema.test_table_15862_c0019_temp_sample_table;
    CREATE TABLE schema.dirty_shared AS
      SELECT explode(array(if(id IS NULL, '1001', NULL), if(name = '', '1002', NULL),
                           if(address = '', '1003', NULL), if(age < 0, '1004', NULL)));
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.FULL,
        merged_rule_groups=((1, 2, 3, 4),),
        isolated_rules=(),
        sampling_percent=50,
        partition_filter="dt=2026-08-04",
    )

    readback = assert_sql_topology(
        sql,
        table_name="test_table_15862_c0019",
        rule_count=4,
        expectation=expectation,
        rule_identities=_rule_identities(4),
    )

    assert readback.effective_scan_count == 1
    assert readback.dirty_targets == ("schema.dirty_shared",)
    assert readback.sampling_enabled is True


def test_sql_topology_rejects_sample_setup_that_scans_source_twice() -> None:
    sql = """
    CREATE TABLE schema.test_table_15862_c0019_temp_sample_table AS
      SELECT * FROM schema.test_table_15862_c0019 WHERE rand() <= 0.5;
    INSERT INTO schema.test_table_15862_c0019_temp_sample_table
      SELECT * FROM schema.test_table_15862_c0019 WHERE rand() <= 0.5;
    INSERT INTO schema.dirty_shared
      SELECT SUM(CASE WHEN id > 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002
      FROM schema.test_table_15862_c0019_temp_sample_table;
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.FULL,
        merged_rule_groups=((1, 2),),
        isolated_rules=(),
        sampling_percent=50,
    )

    with pytest.raises(SqlMergeAssertionError, match="one source setup"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0019",
            rule_count=2,
            expectation=expectation,
            rule_identities=_rule_identities(2),
        )


def test_sql_topology_accepts_partial_groups_and_independent_dirty_targets() -> None:
    sql = """
    INSERT INTO schema.result_rows_merged
      SELECT SUM(CASE WHEN id IS NULL THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002,
             SUM(CASE WHEN age < 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1004,
             SUM(CASE WHEN address = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1005
      FROM schema.test_table_15862_c0005;
    INSERT INTO schema.dirty_rule_1003
      SELECT 1003 AS rule_id FROM schema.test_table_15862_c0005;
    INSERT INTO schema.dirty_rule_1006
      SELECT 1006 AS rule_id FROM schema.test_table_15862_c0005;
    CREATE TABLE schema.dirty_merged AS
      SELECT explode(array(if(id IS NULL, '1001', NULL), if(name = '', '1002', NULL),
                           if(age < 0, '1004', NULL), if(address = '', '1005', NULL)));
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.PARTIAL,
        merged_rule_groups=((1, 2, 4, 5),),
        isolated_rules=(3, 6),
        sampling_percent=None,
    )

    readback = assert_sql_topology(
        sql,
        table_name="test_table_15862_c0005",
        rule_count=6,
        expectation=expectation,
        rule_identities=_rule_identities(6),
    )

    assert readback.effective_scan_count == _PARTIAL_SCAN_GROUP_COUNT
    assert len(readback.dirty_targets) == _PARTIAL_SCAN_GROUP_COUNT


@pytest.mark.parametrize(
    ("sql", "expectation", "message"),
    [
        (
            "INSERT INTO dirty SELECT * FROM test_table_15862_c0019",
            SqlTopologyExpectation(
                mode=MergeMode.FULL,
                merged_rule_groups=((1, 2),),
                isolated_rules=(),
                sampling_percent=50,
            ),
            "sampling",
        ),
        (
            ("INSERT INTO dirty SELECT SUM(CASE WHEN id=1 THEN 1 END) FROM test_table_15862_c0019"),
            SqlTopologyExpectation(
                mode=MergeMode.NONE,
                merged_rule_groups=(),
                isolated_rules=(1, 2),
                sampling_percent=None,
            ),
            "independent",
        ),
    ],
)
def test_sql_topology_rejects_weakened_runtime_sql(
    sql: str,
    expectation: SqlTopologyExpectation,
    message: str,
) -> None:
    with pytest.raises(SqlMergeAssertionError, match=message):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0019",
            rule_count=2,
            expectation=expectation,
            rule_identities=_rule_identities(2),
        )


def test_sql_topology_rejects_wrong_members_with_same_scan_and_target_counts() -> None:
    sql = """
    INSERT INTO schema.dirty_shared
      SELECT SUM(CASE WHEN id IS NULL THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1003
      FROM schema.test_table_15862_c0011;
    INSERT INTO schema.dirty_shared
      SELECT SUM(CASE WHEN address = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002,
             SUM(CASE WHEN age < 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1004
      FROM schema.test_table_15862_c0011;
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.FULL,
        merged_rule_groups=((1, 2), (3, 4)),
        isolated_rules=(),
        sampling_percent=None,
    )

    with pytest.raises(SqlMergeAssertionError, match="rule-group members"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0011",
            rule_count=4,
            expectation=expectation,
            rule_identities=_rule_identities(4),
        )


def test_sql_topology_rejects_reused_target_for_independent_groups() -> None:
    sql = """
    INSERT INTO schema.dirty_reused
      SELECT SUM(CASE WHEN id IS NULL THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002
      FROM schema.test_table_15862_c0005;
    INSERT INTO schema.dirty_reused
      SELECT 1003 AS rule_id FROM schema.test_table_15862_c0005;
    CREATE TABLE schema.dirty_reused AS
      SELECT explode(array(if(id IS NULL, '1001', NULL), if(name = '', '1002', NULL),
                           if(age < 0, '1003', NULL)));
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.PARTIAL,
        merged_rule_groups=((1, 2),),
        isolated_rules=(3,),
        sampling_percent=None,
    )

    with pytest.raises(SqlMergeAssertionError, match="target equivalence"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0005",
            rule_count=3,
            expectation=expectation,
            rule_identities=_rule_identities(3),
        )


def test_sql_topology_allows_sum_case_inside_an_independent_rule() -> None:
    sql = """
    INSERT INTO schema.dirty_rule_1001
      SELECT 1001 AS rule_id, SUM(CASE WHEN id IS NULL THEN 1 ELSE 0 END)
      FROM schema.test_table_15862_c0008;
    INSERT INTO schema.dirty_rule_1002
      SELECT 1002 AS rule_id FROM schema.test_table_15862_c0008;
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.NONE,
        merged_rule_groups=(),
        isolated_rules=(1, 2),
        sampling_percent=None,
    )

    readback = assert_sql_topology(
        sql,
        table_name="test_table_15862_c0008",
        rule_count=2,
        expectation=expectation,
        rule_identities=_rule_identities(2),
    )

    assert readback.sum_case_count == 1


def test_sql_topology_rejects_merged_group_masked_by_another_group_signature() -> None:
    sql = """
    INSERT INTO schema.result_group_1
      SELECT SUM(CASE WHEN id IS NULL THEN 1 ELSE 0 END) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002
      FROM schema.test_table_15862_c0011;
    INSERT INTO schema.result_group_2
      SELECT count(address) AS hit_cnt_rule_1003,
             count(age) AS hit_cnt_rule_1004
      FROM schema.test_table_15862_c0011;
    CREATE TABLE schema.dirty_group_1 AS
      SELECT explode(array(if(id IS NULL, '1001', NULL), if(name = '', '1002', NULL)));
    CREATE TABLE schema.dirty_group_2 AS
      SELECT explode(array(if(address = '', '1003', NULL), if(age < 0, '1004', NULL)));
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.FULL,
        merged_rule_groups=((1, 2), (3, 4)),
        isolated_rules=(),
        sampling_percent=None,
    )

    with pytest.raises(SqlMergeAssertionError, match="own SUM"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0011",
            rule_count=4,
            expectation=expectation,
            rule_identities=_rule_identities(4),
        )


def test_sql_topology_binds_distinct_to_the_declared_group() -> None:
    sql = """
    INSERT INTO schema.result_group_1
      SELECT count(DISTINCT id) AS hit_cnt_rule_1001,
             SUM(CASE WHEN name = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1002
      FROM schema.test_table_15862_c0003;
    INSERT INTO schema.result_group_2
      SELECT SUM(CASE WHEN address = '' THEN 1 ELSE 0 END) AS hit_cnt_rule_1003,
             SUM(CASE WHEN age < 0 THEN 1 ELSE 0 END) AS hit_cnt_rule_1004
      FROM schema.test_table_15862_c0003;
    CREATE TABLE schema.dirty_group_1 AS
      SELECT explode(array(if(id IS NULL, '1001', NULL), if(name = '', '1002', NULL)));
    CREATE TABLE schema.dirty_group_2 AS
      SELECT explode(array(if(address = '', '1003', NULL), if(age < 0, '1004', NULL)));
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.FULL,
        merged_rule_groups=((1, 2), (3, 4)),
        isolated_rules=(),
        sampling_percent=None,
        distinct_rule_groups=((3, 4),),
    )

    with pytest.raises(SqlMergeAssertionError, match="declared SQL rule group"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0003",
            rule_count=4,
            expectation=expectation,
            rule_identities=_rule_identities(4),
        )


def test_sql_topology_rejects_numeric_threshold_as_rule_identity() -> None:
    sql = """
    INSERT INTO schema.dirty_rule_1002
      SELECT 1002 AS rule_id FROM schema.test_table_15862_c0008 WHERE id > 1001;
    INSERT INTO schema.dirty_rule_1003
      SELECT 1003 AS rule_id FROM schema.test_table_15862_c0008;
    """
    expectation = SqlTopologyExpectation(
        mode=MergeMode.NONE,
        merged_rule_groups=(),
        isolated_rules=(1, 2, 3),
        sampling_percent=None,
    )

    with pytest.raises(SqlMergeAssertionError, match="preserve every declared scan group"):
        assert_sql_topology(
            sql,
            table_name="test_table_15862_c0008",
            rule_count=3,
            expectation=expectation,
            rule_identities=_rule_identities(3),
        )
