from __future__ import annotations

from typing import TYPE_CHECKING
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from data_assets_playwright_web_ui.domains.data_quality.json_value_validation import (
    JsonValueAssertions,
    JsonValueValidationContractError,
    parse_sampling_readback,
)
from data_assets_playwright_web_ui.domains.data_quality.json_value_validation.assertions import (
    canonical_quality_report_header_indexes,
)

if TYPE_CHECKING:
    from pathlib import Path


def _failed_detail_workbook(
    path: Path,
    *,
    payload_color: str = "FFFF0000",
    id_has_fill: bool = False,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(("id", "payload", "name"))
    sheet.append((2, '{"product":{"code":"invalid_code","price":"abc"}}', "invalid"))
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor=payload_color)
    if id_has_fill:
        sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    workbook.save(path)
    workbook.close()


def test_failed_detail_workbook_requires_exact_red_and_default_other_fills(
    tmp_path: Path,
) -> None:
    target = tmp_path / "failed_detail.xlsx"
    _failed_detail_workbook(target)

    JsonValueAssertions().inspect_failed_detail_workbook(target)


def test_failed_detail_workbook_rejects_imprecise_red(tmp_path: Path) -> None:
    target = tmp_path / "failed_detail.xlsx"
    _failed_detail_workbook(target, payload_color="FFFF0001")

    with pytest.raises(JsonValueValidationContractError, match="标准红色"):
        JsonValueAssertions().inspect_failed_detail_workbook(target)


def test_failed_detail_workbook_rejects_fill_on_non_validated_field(
    tmp_path: Path,
) -> None:
    target = tmp_path / "failed_detail.xlsx"
    _failed_detail_workbook(target, id_has_fill=True)

    with pytest.raises(JsonValueValidationContractError, match="默认无填充"):
        JsonValueAssertions().inspect_failed_detail_workbook(target)


def test_failed_detail_workbook_rejects_excessive_zip_compression_ratio(
    tmp_path: Path,
) -> None:
    target = tmp_path / "suspicious.xlsx"
    with ZipFile(target, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x" * 200_000)
        archive.writestr("xl/workbook.xml", "<workbook />")

    with pytest.raises(JsonValueValidationContractError, match="压缩比"):
        JsonValueAssertions().inspect_failed_detail_workbook(target)


def test_rule_library_reader_reuses_safe_archive_preflight(tmp_path: Path) -> None:
    target = tmp_path / "suspicious_rule_library.xlsx"
    with ZipFile(target, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x" * 200_000)
        archive.writestr("xl/workbook.xml", "<workbook />")

    with pytest.raises(JsonValueValidationContractError, match="压缩比"):
        JsonValueAssertions().inspect_exported_rule_library(target)


def test_rule_library_reader_bounds_worksheet_columns(tmp_path: Path) -> None:
    target = tmp_path / "oversized_rule_library.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.cell(row=1, column=257, value="outside-boundary")
    workbook.save(target)
    workbook.close()

    with pytest.raises(JsonValueValidationContractError, match="列数"):
        JsonValueAssertions().inspect_exported_rule_library(target)


def test_rule_library_reader_accepts_bounded_complete_export(tmp_path: Path) -> None:
    target = tmp_path / "rule_library.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        (
            "格式-json格式校验",
            "有效性校验",
            "字段",
            "校验json类型的字段中key对应的value值是否符合规范要求",
        ),
    )
    workbook.save(target)
    workbook.close()

    JsonValueAssertions().inspect_exported_rule_library(target)


def test_quality_report_headers_require_exact_canonical_seven_columns() -> None:
    headers = (
        "规则类型",
        "规则名称",
        "字段类型",
        "质检结果",
        "未通过原因",
        "详情说明",
        "操作",
    )

    assert canonical_quality_report_header_indexes(headers) == {
        header: index for index, header in enumerate(headers)
    }


def test_quality_report_headers_reject_task_query_rule_view_descriptions() -> None:
    rule_view_labels = (
        "字段",
        "统计函数",
        "过滤条件",
        "校验方法",
        "期望值",
        "规则强弱",
        "规则描述",
    )

    with pytest.raises(JsonValueValidationContractError, match="canonical 七列表格契约"):
        canonical_quality_report_header_indexes(rule_view_labels)


def test_sampling_parser_returns_actual_visible_values_with_or_without_about() -> None:
    assert parse_sampling_readback("抽样比例 50%\uff0c参与校验数据量约 10 条").as_json() == {
        "ratio_percent": 50,
        "validated_count": 10,
    }
    assert parse_sampling_readback("抽样比例 50%\uff0c参与校验数据量 10 条").as_json() == {
        "ratio_percent": 50,
        "validated_count": 10,
    }


@pytest.mark.parametrize(
    "text",
    [
        "抽样比例 101%\uff0c参与校验数据量 10 条",
        "抽样比例 50%\uff0c参与校验数据量 10000001 条",
        "抽样比例 50%\uff0c参与校验数据量 10 条\uff0c其他数据量 11 条",
    ],
)
def test_sampling_parser_rejects_out_of_bounds_or_ambiguous_values(text: str) -> None:
    with pytest.raises(JsonValueValidationContractError):
        parse_sampling_readback(text)
