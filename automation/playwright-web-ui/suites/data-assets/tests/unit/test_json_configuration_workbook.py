from __future__ import annotations

from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from data_assets_playwright_web_ui.domains.data_quality.json_configuration.model import (
    JsonImportRow,
)
from data_assets_playwright_web_ui.domains.data_quality.json_configuration.workbook import (
    IMPORT_SHEET_NAMES,
    JsonConfigurationWorkbook,
    WorkbookContractError,
)

if TYPE_CHECKING:
    from collections.abc import Callable
    from pathlib import Path


def _inspect_template(path: Path) -> object:
    return JsonConfigurationWorkbook.inspect_template(path)


def _inspect_export(path: Path) -> object:
    return JsonConfigurationWorkbook.inspect_export(path)


def _inspect_error_value(path: Path) -> object:
    return JsonConfigurationWorkbook.inspect_error_cell(
        path,
        sheet_name="一层",
        value="alpha",
    )


def _inspect_error_coordinate(path: Path) -> object:
    return JsonConfigurationWorkbook.inspect_error_coordinate(
        path,
        sheet_name="一层",
        coordinate="A2",
    )


def test_builder_writes_the_five_level_import_contract(tmp_path: Path) -> None:
    target = tmp_path / "json_format_import_15696.xlsx"
    JsonConfigurationWorkbook.build_import(
        target,
        rows=(
            JsonImportRow(level=1, parents=(), key="root", chinese_name="根", value_format=""),
            JsonImportRow(
                level=2,
                parents=("root",),
                key="child",
                chinese_name="子",
                value_format=r"^\d+$",
            ),
        ),
    )

    workbook = load_workbook(target, read_only=True, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == IMPORT_SHEET_NAMES
        assert tuple(cell.value for cell in workbook["一层"][1]) == (
            "* key",
            "中文名称",
            "value格式",
        )
        assert tuple(cell.value for cell in workbook["二层"][1]) == (
            "* 第一层级key名",
            "* key",
            "中文名称",
            "value格式",
        )
        assert tuple(cell.value for cell in workbook["二层"][2]) == (
            "root",
            "child",
            "子",
            r"^\d+$",
        )
    finally:
        workbook.close()


def test_template_inspection_requires_exact_sheet_and_header_order(tmp_path: Path) -> None:
    target = tmp_path / "template.xlsx"
    JsonConfigurationWorkbook.build_import(target, rows=())

    inspection = JsonConfigurationWorkbook.inspect_template(target)

    assert inspection.sheet_names == IMPORT_SHEET_NAMES
    assert inspection.headers[4] == (
        "* 第一层级key名",
        "* 第二层级key名",
        "* 第三层级key名",
        "* 第四层级key名",
        "* key",
        "中文名称",
        "value格式",
    )


def test_error_cell_inspection_reports_red_fill_and_comment(tmp_path: Path) -> None:
    target = tmp_path / "json_format_error_20260809.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "二层"
    sheet.append(["* 第一层级key名", "* key", "中文名称", "value格式"])
    sheet.append(["missing-parent", "orphan", "孤儿", ""])
    parent = sheet.cell(row=2, column=1)
    parent.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    parent.comment = Comment("上一层级无相同key名匹配", "system")
    workbook.save(target)

    inspection = JsonConfigurationWorkbook.inspect_error_cell(
        target,
        sheet_name="二层",
        value="missing-parent",
    )

    assert inspection.coordinate == "A2"
    assert inspection.is_red
    assert inspection.comment == "上一层级无相同key名匹配"


def test_export_reader_returns_named_rows_without_blank_records(tmp_path: Path) -> None:
    target = tmp_path / "json_format_20260809.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "配置"
    sheet.append(["key", "中文名称", "value 格式", "数据源类型", "创建人"])
    sheet.append(["alpha", "甲", r"^[a-z]+$", "Hive2.x", "admin"])
    sheet.append([None, None, None, None, None])
    workbook.save(target)

    rows = JsonConfigurationWorkbook.read_export(target)

    assert rows == (
        {
            "key": "alpha",
            "中文名称": "甲",
            "value 格式": r"^[a-z]+$",
            "数据源类型": "Hive2.x",
            "创建人": "admin",
        },
    )


def test_export_inspection_preserves_the_exact_ordered_header_contract(tmp_path: Path) -> None:
    target = tmp_path / "json_format_20260809.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["key", "中文名称", "value 格式"])
    sheet.append(["alpha", "甲", r"^[a-z]+$"])
    workbook.save(target)

    inspection = JsonConfigurationWorkbook.inspect_export(target)

    assert inspection.headers == ("key", "中文名称", "value 格式")
    assert inspection.rows == ({"key": "alpha", "中文名称": "甲", "value 格式": r"^[a-z]+$"},)


@pytest.mark.parametrize(
    "reader",
    [
        JsonConfigurationWorkbook.inspect_template,
        JsonConfigurationWorkbook.read_export,
    ],
)
def test_downloaded_workbook_readers_reject_oversized_files_before_parsing(
    tmp_path: Path,
    reader: Callable[[Path], object],
) -> None:
    target = tmp_path / "oversized.xlsx"
    with target.open("wb") as stream:
        stream.seek(20 * 1024 * 1024)
        stream.write(b"x")

    with pytest.raises(
        WorkbookContractError,
        match="workbook exceeds the safe file-size limit",
    ):
        reader(target)


@pytest.mark.parametrize(
    "reader",
    [
        _inspect_template,
        _inspect_export,
        _inspect_error_value,
        _inspect_error_coordinate,
    ],
)
def test_every_workbook_reader_closes_openpyxl_resource(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    reader: Callable[[Path], object],
) -> None:
    target = JsonConfigurationWorkbook.build_import(
        tmp_path / "downloaded.xlsx",
        rows=(JsonImportRow(1, (), "alpha", "甲", ""),),
    )
    original_close = Workbook.close
    closed: list[Workbook] = []

    def tracked_close(workbook: Workbook) -> None:
        closed.append(workbook)
        original_close(workbook)

    monkeypatch.setattr(Workbook, "close", tracked_close)

    reader(target)

    assert len(closed) == 1
