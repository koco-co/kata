"""Strong UI and downloaded-workbook assertions for JSON value validation."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import PurePosixPath
from typing import TYPE_CHECKING, Final
from zipfile import BadZipFile, ZipFile

from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from playwright.sync_api import expect

from data_assets_playwright_web_ui.domains.data_quality.json_value_validation.model import (
    CUSTOM_REGEX_RULE,
    JSON_FORMAT_RULE,
    SamplingReadback,
)

if TYPE_CHECKING:
    from pathlib import Path
    from zipfile import ZipInfo

    from openpyxl.workbook.workbook import Workbook
    from playwright.sync_api import Locator

_UI_TIMEOUT_MS: Final = 30_000
_LONG_TIMEOUT_MS: Final = 600_000
_EXPECTED_INVALID_RECORD_ID: Final = 2
_STANDARD_RED_RGB: Final = "FFFF0000"
_MAX_XLSX_FILE_BYTES: Final = 20 * 1024 * 1024
_MAX_XLSX_EXPANDED_BYTES: Final = 100 * 1024 * 1024
_MAX_XLSX_MEMBER_BYTES: Final = 50 * 1024 * 1024
_MAX_XLSX_ENTRIES: Final = 1_024
_MAX_XLSX_COMPRESSION_RATIO: Final = 100
_MAX_WORKSHEETS: Final = 16
_MAX_WORKSHEET_ROWS: Final = 100_000
_MAX_WORKSHEET_COLUMNS: Final = 256
_MAX_CELL_CHARACTERS: Final = 32_767
_EXPECTED_DETAIL_ROWS: Final = 2
_EXPECTED_DETAIL_COLUMNS: Final = 3
_CANONICAL_QUALITY_REPORT_HEADERS: Final = (
    "规则类型",
    "规则名称",
    "字段类型",
    "质检结果",
    "未通过原因",
    "详情说明",
    "操作",
)


class JsonValueValidationContractError(AssertionError):
    """Raised when rendered business state violates the canonical feature contract."""


def canonical_quality_report_header_indexes(headers: tuple[str, ...]) -> dict[str, int]:
    """Require and index the exact canonical seven-column quality-report contract."""
    normalized = tuple(" ".join(header.split()) for header in headers)
    if normalized != _CANONICAL_QUALITY_REPORT_HEADERS:
        message = (
            "目标任务实例的 canonical 七列表格契约不满足。"
            f"要求={_CANONICAL_QUALITY_REPORT_HEADERS!r}, 实际={normalized!r}"
        )
        raise JsonValueValidationContractError(message)
    return {header: index for index, header in enumerate(normalized)}


def parse_sampling_readback(text: str) -> SamplingReadback:
    """Parse unique bounded sampling values from visible instance-detail text."""
    ratio_values = {int(value) for value in re.findall(r"(?<!\d)(\d{1,4})\s*%(?!\d)", text)}
    count_values = {int(value) for value in re.findall(r"(?<!\d)(?:约\s*)?(\d{1,10})\s*条", text)}
    if len(ratio_values) != 1 or len(count_values) != 1:
        message = (
            "实例详情必须唯一展示抽样比例和约参与校验条数。"
            f"实际比例={sorted(ratio_values)}, 条数={sorted(count_values)}"
        )
        raise JsonValueValidationContractError(message)
    try:
        return SamplingReadback(
            ratio_percent=next(iter(ratio_values)),
            validated_count=next(iter(count_values)),
        )
    except ValueError as error:
        message = "实例详情抽样数值超出业务安全边界"
        raise JsonValueValidationContractError(message) from error


@dataclass(frozen=True, slots=True)
class JsonValueAssertions:
    """Assert user-visible JSON value-format behavior without permissive fallbacks."""

    def expect_rule_order(self, dropdown: Locator) -> None:
        """Assert JSON validation is immediately before custom-regex validation."""
        options = dropdown.locator(".ant-select-item-option")
        labels = tuple(text.strip() for text in options.all_inner_texts() if text.strip())
        try:
            json_index = labels.index(JSON_FORMAT_RULE)
            regex_index = labels.index(CUSTOM_REGEX_RULE)
        except ValueError as error:
            message = "统计规则下拉必须同时展示 JSON 格式校验与自定义正则"
            raise JsonValueValidationContractError(message) from error
        if json_index + 1 != regex_index:
            message = "格式-json格式校验必须紧邻并位于格式校验-自定义正则上方"
            raise JsonValueValidationContractError(message)

    def expect_rule_option(self, dropdown: Locator, *, visible: bool) -> None:
        """Assert the exact JSON validation option is present or absent."""
        option = dropdown.get_by_text(JSON_FORMAT_RULE, exact=True)
        if visible:
            expect(option, f"统计规则下拉必须展示“{JSON_FORMAT_RULE}”").to_be_visible(
                timeout=_UI_TIMEOUT_MS,
            )
            expect(option).not_to_have_attribute("aria-disabled", "true")
        else:
            expect(option, f"统计规则下拉不得展示“{JSON_FORMAT_RULE}”").to_have_count(0)

    def expect_no_independent_help(self, function_row: Locator) -> None:
        """Assert the JSON function row has no independent help or tooltip trigger."""
        expect(
            function_row.locator(
                ".anticon-question-circle, .anticon-info-circle, [aria-describedby]",
            ),
            "JSON 格式校验统计函数行不得展示独立提示入口",
        ).to_have_count(0)

    def expect_key_state(
        self,
        node: Locator,
        *,
        checked: bool,
        disabled: bool,
    ) -> None:
        """Assert AntD TreeSelect exposes exact checked and disabled state."""
        checkbox = node.locator(".ant-select-tree-checkbox")
        expect(checkbox).to_be_visible()
        checked_pattern = re.compile(r"(?:^|\s)ant-select-tree-checkbox-checked(?:\s|$)")
        disabled_pattern = re.compile(r"(?:^|\s)ant-select-tree-treenode-disabled(?:\s|$)")
        if checked:
            expect(checkbox).to_have_class(checked_pattern)
        else:
            expect(checkbox).not_to_have_class(checked_pattern)
        if disabled:
            expect(node).to_have_class(disabled_pattern)
        else:
            expect(node).not_to_have_class(disabled_pattern)

    def expect_selected_tags(self, selector: Locator, expected: tuple[str, ...]) -> None:
        """Assert exact selected key tags, excluding AntD overflow counters."""
        tags = selector.locator(".ant-select-selection-item")
        actual = tuple(text.strip() for text in tags.all_inner_texts() if text.strip())
        if actual != expected:
            message = f"校验 key 回显应为 {expected!r}。实际为 {actual!r}"
            raise JsonValueValidationContractError(message)

    def expect_saved_rule(
        self,
        rule_form: Locator,
        *,
        field: str,
        keys: tuple[str, ...],
        description: str | None = None,
    ) -> None:
        """Assert all persisted rule parameters from the reopened editor."""
        for text in ("字段级", field, JSON_FORMAT_RULE, *keys, "强规则"):
            expect(rule_form, f"已保存规则必须回显“{text}”").to_contain_text(text)
        if description is not None:
            expect(rule_form).to_contain_text(description)

    def expect_task_result(  # noqa: PLR0913
        self,
        detail: Locator,
        *,
        result: str,
        keys: tuple[str, ...],
        field_type: str | None = None,
        failure_reason: str | None = None,
        detail_text: str | None = None,
        has_detail: bool | None = None,
    ) -> None:
        """Assert the rule row has complete business result semantics."""
        row = detail.locator(".ruleView").filter(has_text=JSON_FORMAT_RULE).first
        expect(row, "实例详情必须展示 JSON 格式校验规则卡片").to_be_visible(
            timeout=_LONG_TIMEOUT_MS,
        )
        for text in ("有效性校验", JSON_FORMAT_RULE, result, *keys):
            expect(row).to_contain_text(text)
        if field_type is not None:
            expect(row).to_contain_text(field_type)
        if failure_reason is not None:
            expect(row).to_contain_text(failure_reason)
        if detail_text is not None:
            expect(row).to_contain_text(detail_text)
        detail_action = row.get_by_text(re.compile(r"查看(?:明细|详情)"))
        if has_detail is True:
            expect(detail_action).to_be_visible()
        elif has_detail is False:
            expect(detail_action).to_have_count(0)

    def expect_dirty_ids(
        self,
        drawer: Locator,
        *,
        present: tuple[int, ...],
        absent: tuple[int, ...],
        highlighted_field: str,
    ) -> None:
        """Assert exact dirty record IDs and the highlighted validated field."""
        table = drawer.locator(".ant-table")
        expect(table, "校验明细必须展示数据表").to_be_visible(timeout=_UI_TIMEOUT_MS)
        for record_id in present:
            expect(
                table.locator("tbody tr").filter(
                    has=table.get_by_text(str(record_id), exact=True),
                ),
            ).to_be_visible()
        for record_id in absent:
            expect(
                table.locator("tbody tr").filter(
                    has=table.get_by_text(str(record_id), exact=True),
                ),
            ).to_have_count(0)
        header = table.locator("thead th").filter(has_text=highlighted_field).first
        expect(header, f"明细必须展示“{highlighted_field}”列").to_be_visible()
        column_index = header.evaluate("element => element.cellIndex")
        if not isinstance(column_index, int):
            message = f"无法读取“{highlighted_field}”列位置"
            raise JsonValueValidationContractError(message)
        highlighted_cells = table.locator(
            f"tbody tr td:nth-child({column_index + 1}) [style*='color'], "
            f"tbody tr td:nth-child({column_index + 1}).error, "
            f"tbody tr td:nth-child({column_index + 1}) .error",
        )
        expect(highlighted_cells.first, f"“{highlighted_field}”异常单元格必须标红").to_be_visible()

    def expect_dirty_ids_within(
        self,
        drawer: Locator,
        *,
        allowed: frozenset[int],
        highlighted_field: str,
    ) -> tuple[int, ...]:
        """Require a non-empty dirty set containing only dynamically sampled invalid IDs."""
        table = drawer.locator(".ant-table")
        expect(table, "抽样不通过时必须展示脏数据表").to_be_visible(
            timeout=_UI_TIMEOUT_MS,
        )
        id_header = (
            table.locator("thead th")
            .filter(
                has_text=re.compile(r"^\s*id\s*$"),
            )
            .first
        )
        expect(id_header, "抽样明细必须展示 id 列").to_be_visible()
        column_index = id_header.evaluate("element => element.cellIndex")
        if not isinstance(column_index, int):
            message = "无法读取抽样明细 id 列位置"
            raise JsonValueValidationContractError(message)
        raw_ids = table.locator(
            f"tbody tr td:nth-child({column_index + 1})",
        ).all_inner_texts()
        try:
            actual = tuple(int(value.strip()) for value in raw_ids if value.strip())
        except ValueError as error:
            message = "抽样明细 id 必须为整数"
            raise JsonValueValidationContractError(message) from error
        if not actual or not set(actual).issubset(allowed):
            message = f"抽样脏数据必须非空且仅来自 {sorted(allowed)}。实际为 {actual}"
            raise JsonValueValidationContractError(message)
        header = table.locator("thead th").filter(has_text=highlighted_field).first
        expect(header, f"抽样明细必须展示“{highlighted_field}”列").to_be_visible()
        highlighted_index = header.evaluate("element => element.cellIndex")
        if not isinstance(highlighted_index, int):
            message = f"无法读取“{highlighted_field}”列位置"
            raise JsonValueValidationContractError(message)
        highlighted_cells = table.locator(
            f"tbody tr td:nth-child({highlighted_index + 1}) [style*='color'], "
            f"tbody tr td:nth-child({highlighted_index + 1}).error, "
            f"tbody tr td:nth-child({highlighted_index + 1}) .error",
        )
        expect(
            highlighted_cells.first,
            f"“{highlighted_field}”抽样异常单元格必须标红",
        ).to_be_visible()
        return actual

    def expect_dirty_field_value(
        self,
        drawer: Locator,
        *,
        record_id: int,
        field: str,
        invalid_value: str,
    ) -> str:
        """Assert one exact dirty row visibly contains the canonical invalid value."""
        table = drawer.locator(".ant-table")
        id_index = self._table_column_index(table, "id")
        field_index = self._table_column_index(table, field)
        matching_rows = tuple(
            row
            for row in self._table_rows(table)
            if self._cell_text(row, id_index) == str(record_id)
        )
        if len(matching_rows) != 1:
            message = f"脏数据明细必须唯一展示 id={record_id}。实际匹配 {len(matching_rows)} 行"
            raise JsonValueValidationContractError(message)
        actual = self._cell_text(matching_rows[0], field_index)
        if invalid_value not in actual:
            message = (
                f"id={record_id} 的 {field} 必须展示不合规值“{invalid_value}”。实际为 {actual!r}"
            )
            raise JsonValueValidationContractError(message)
        return actual

    def expect_exact_dirty_ids(
        self,
        drawer: Locator,
        *,
        expected: frozenset[int],
        highlighted_field: str,
    ) -> tuple[int, ...]:
        """Prove the complete unpaginated dirty-ID set and highlighted field."""
        table = drawer.locator(".ant-table")
        expect(table, "校验明细必须展示数据表").to_be_visible(timeout=_UI_TIMEOUT_MS)
        id_index = self._table_column_index(table, "id")
        rows = self._table_rows(table)
        raw_ids = tuple(self._cell_text(row, id_index) for row in rows)
        try:
            actual = tuple(int(value) for value in raw_ids)
        except ValueError as error:
            message = f"脏数据 id 必须全部为整数。实际为 {raw_ids!r}"
            raise JsonValueValidationContractError(message) from error
        if len(actual) != len(set(actual)) or frozenset(actual) != expected:
            message = f"脏数据 ID 集必须精确为 {sorted(expected)}。实际为 {actual}"
            raise JsonValueValidationContractError(message)
        expect(
            drawer.get_by_text(
                re.compile(rf"^共\s*{len(expected)}\s*条问题数据$"),
            ),
            "明细总数必须与完整脏数据 ID 集一致",
        ).to_be_visible(timeout=_UI_TIMEOUT_MS)
        expect(
            drawer.locator(".ant-pagination"),
            "明细表禁用分页时才可由当前行证明完整 ID 集",
        ).to_have_count(0)
        highlighted_index = self._table_column_index(table, highlighted_field)
        highlighted_cells = table.locator(
            f"tbody tr td:nth-child({highlighted_index + 1}) [style*='color'], "
            f"tbody tr td:nth-child({highlighted_index + 1}).error, "
            f"tbody tr td:nth-child({highlighted_index + 1}) .error",
        )
        expect(
            highlighted_cells,
            f"每条“{highlighted_field}”异常值必须标红",
        ).to_have_count(len(expected))
        return actual

    def expect_sampling_readback(
        self,
        detail: Locator,
        *,
        expected_ratio_percent: int,
        expected_validated_count: int,
    ) -> SamplingReadback:
        """Parse, bound, and assert actual sampling values rendered by the report."""
        readback = parse_sampling_readback(detail.inner_text())
        if readback.ratio_percent != expected_ratio_percent:
            message = f"抽样比例必须为 {expected_ratio_percent}%。实际为 {readback.ratio_percent}%"
            raise JsonValueValidationContractError(message)
        if readback.validated_count != expected_validated_count:
            message = (
                f"参与校验数量必须为 {expected_validated_count}。实际为 {readback.validated_count}"
            )
            raise JsonValueValidationContractError(message)
        return readback

    def expect_quality_report_rule_table(  # noqa: PLR0913
        self,
        detail: Locator,
        *,
        field_type: str,
        result: str,
        reason: str,
        detail_text: str,
        has_detail: bool,
    ) -> None:
        """Assert the canonical seven-column JSON row in the linked task detail."""
        table, indexes = self._canonical_quality_report_table(detail)
        rows = self._table_rows(table)
        matching_rows = tuple(
            row
            for row in rows
            if self._cell_text(row, indexes["规则类型"]) == "有效性校验"
            and self._cell_text(row, indexes["规则名称"]) == JSON_FORMAT_RULE
        )
        if len(matching_rows) != 1:
            message = (
                "目标任务实例的 canonical 七列表格必须唯一展示"
                f"“有效性校验/{JSON_FORMAT_RULE}”规则行。实际匹配 {len(matching_rows)} 行"
            )
            raise JsonValueValidationContractError(message)
        row = matching_rows[0]
        if row.locator("td").count() != len(_CANONICAL_QUALITY_REPORT_HEADERS):
            message = "目标 JSON 规则行的单元格数量必须与 canonical 七列表头一致"
            raise JsonValueValidationContractError(message)
        expected = {
            "规则类型": "有效性校验",
            "规则名称": JSON_FORMAT_RULE,
            "字段类型": field_type,
            "质检结果": result,
            "未通过原因": reason,
            "详情说明": detail_text,
            "操作": "查看详情" if has_detail else "",
        }
        for header, value in expected.items():
            actual = self._cell_text(row, indexes[header])
            if actual != value:
                message = f"质量报告“{header}”列必须精确为 {value!r}。实际为 {actual!r}"
                raise JsonValueValidationContractError(message)

    def inspect_exported_rule_library(self, path: Path) -> None:
        """Assert the downloaded workbook contains the complete built-in rule record."""
        values = _workbook_text(path)
        for expected in (
            JSON_FORMAT_RULE,
            "有效性校验",
            "字段",
            "校验json类型的字段中key对应的value值是否符合规范要求",
        ):
            if expected not in values:
                message = f"导出规则库必须包含“{expected}”"
                raise JsonValueValidationContractError(message)

    def inspect_failed_detail_workbook(self, path: Path) -> None:
        """Assert the downloaded dirty-data workbook marks only the validated field."""
        _require_safe_xlsx(path)
        try:
            workbook = load_workbook(path, data_only=True, keep_links=False)
        except (OSError, ValueError, DefusedXmlException) as error:
            message = "下载明细必须是可读取的 xlsx 文件"
            raise JsonValueValidationContractError(message) from error
        try:
            worksheet = _bounded_active_worksheet(workbook)
            _assert_failed_detail_content(worksheet)
        finally:
            workbook.close()

    @staticmethod
    def _table_rows(table: Locator) -> tuple[Locator, ...]:
        rows = table.locator("tbody tr.ant-table-row")
        return tuple(rows.nth(index) for index in range(rows.count()))

    @staticmethod
    def _table_column_index(table: Locator, header: str) -> int:
        locator = table.locator("thead th").filter(
            has_text=re.compile(rf"^\s*{re.escape(header)}\s*$"),
        )
        if locator.count() != 1:
            message = f"数据表必须唯一展示“{header}”列"
            raise JsonValueValidationContractError(message)
        index = locator.evaluate("element => element.cellIndex")
        if not isinstance(index, int):
            message = f"无法读取数据表“{header}”列位置"
            raise JsonValueValidationContractError(message)
        return index

    @staticmethod
    def _cell_text(row: Locator, index: int) -> str:
        return " ".join(row.locator("td").nth(index).inner_text().split())

    def _canonical_quality_report_table(
        self,
        detail: Locator,
    ) -> tuple[Locator, dict[str, int]]:
        # 当前 taskQuery 的 RuleView 是描述卡片; 这里只接受 canonical 明示的七列表格。
        tables = detail.locator(".ant-table")
        observed: list[tuple[str, ...]] = []
        matches: list[tuple[Locator, dict[str, int]]] = []
        for offset in range(tables.count()):
            table = tables.nth(offset)
            if not table.is_visible():
                continue
            headers = tuple(table.locator("thead th").all_inner_texts())
            observed.append(tuple(" ".join(header.split()) for header in headers))
            try:
                indexes = canonical_quality_report_header_indexes(headers)
            except JsonValueValidationContractError:
                continue
            matches.append((table, indexes))
        if len(matches) != 1:
            message = (
                "目标任务实例必须唯一展示 canonical 七列表格; "
                f"实际匹配 {len(matches)} 张, 可见表头={observed!r}"
            )
            raise JsonValueValidationContractError(message)
        return matches[0]


def _is_exact_standard_red_fill(cell: object) -> bool:
    fill = getattr(cell, "fill", None)
    foreground = getattr(fill, "fgColor", None)
    fill_type = getattr(fill, "fill_type", None)
    color_type = getattr(foreground, "type", None)
    rgb = getattr(foreground, "rgb", None)
    return (
        fill_type == "solid"
        and color_type == "rgb"
        and isinstance(rgb, str)
        and rgb.upper() == _STANDARD_RED_RGB
    )


def _bounded_active_worksheet(workbook: Workbook) -> Worksheet:
    worksheets = _bounded_worksheets(workbook)
    worksheet = workbook.active
    if not isinstance(worksheet, Worksheet) or worksheet not in worksheets:
        message = "下载明细 xlsx 必须包含活动工作表"
        raise JsonValueValidationContractError(message)
    _worksheet_nonempty_texts(worksheet)
    return worksheet


def _bounded_worksheets(workbook: Workbook) -> tuple[Worksheet, ...]:
    if not 0 < len(workbook.sheetnames) <= _MAX_WORKSHEETS:
        message = "xlsx 工作表数量超出安全边界"
        raise JsonValueValidationContractError(message)
    worksheets = tuple(workbook.worksheets)
    if len(worksheets) != len(workbook.sheetnames):
        message = "xlsx 仅允许普通工作表"
        raise JsonValueValidationContractError(message)
    for worksheet in worksheets:
        _require_worksheet_dimensions(worksheet)
    return worksheets


def _require_worksheet_dimensions(worksheet: Worksheet) -> None:
    if not 1 <= worksheet.max_row <= _MAX_WORKSHEET_ROWS:
        message = "xlsx 工作表行数超出安全边界"
        raise JsonValueValidationContractError(message)
    if not 1 <= worksheet.max_column <= _MAX_WORKSHEET_COLUMNS:
        message = "xlsx 工作表列数超出安全边界"
        raise JsonValueValidationContractError(message)


def _worksheet_nonempty_texts(worksheet: Worksheet) -> tuple[str, ...]:
    values: list[str] = []
    for cells in worksheet.iter_rows():
        for cell in cells:
            text = "" if cell.value is None else str(cell.value).strip()
            if len(text) > _MAX_CELL_CHARACTERS:
                message = f"xlsx 单元格 {cell.coordinate} 超出文本长度上限"
                raise JsonValueValidationContractError(message)
            if text:
                values.append(text)
    return tuple(values)


def _assert_failed_detail_content(worksheet: Worksheet) -> None:
    headers = tuple(str(cell.value or "").strip() for cell in worksheet[1])
    if headers != ("id", "payload", "name"):
        message = f"明细列必须精确为 id、payload、name。实际为 {headers}"
        raise JsonValueValidationContractError(message)
    if (
        worksheet.max_row != _EXPECTED_DETAIL_ROWS
        or worksheet.max_column != _EXPECTED_DETAIL_COLUMNS
    ):
        message = "下载明细必须仅包含表头及 id=2 的完整三列脏数据"
        raise JsonValueValidationContractError(message)
    headers_by_name = {header: index for index, header in enumerate(headers, start=1)}
    target_row = _EXPECTED_DETAIL_ROWS
    actual_id = str(worksheet.cell(target_row, headers_by_name["id"]).value).strip()
    if actual_id != str(_EXPECTED_INVALID_RECORD_ID):
        message = "下载明细必须包含且仅包含 id=2 的不合规记录"
        raise JsonValueValidationContractError(message)
    payload = worksheet.cell(target_row, headers_by_name["payload"])
    if not _is_exact_standard_red_fill(payload):
        message = "payload 校验字段必须使用 solid/FFFF0000 标准红色填充"
        raise JsonValueValidationContractError(message)
    for field in ("id", "name"):
        cell = worksheet.cell(target_row, headers_by_name[field])
        if getattr(cell.fill, "fill_type", None) is not None:
            column = get_column_letter(headers_by_name[field])
            message = f"非校验字段 {column} 必须保持默认无填充"
            raise JsonValueValidationContractError(message)


def _require_safe_xlsx(path: Path) -> None:
    """Reject oversized, unsafe, or suspicious XLSX archives before XML parsing."""
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        message = "下载明细必须是本地普通 .xlsx 文件"
        raise JsonValueValidationContractError(message)
    try:
        file_size = path.stat().st_size
    except OSError as error:
        message = "下载明细文件不可读取"
        raise JsonValueValidationContractError(message) from error
    if not 0 < file_size <= _MAX_XLSX_FILE_BYTES:
        message = "下载明细文件大小超出安全边界"
        raise JsonValueValidationContractError(message)
    entries = _xlsx_entries(path)
    _validate_xlsx_entry_index(entries)
    expanded_total = 0
    for entry in entries:
        expanded_total = _validated_expanded_total(entry, expanded_total)


def _xlsx_entries(path: Path) -> tuple[ZipInfo, ...]:
    try:
        with ZipFile(path) as archive:
            return tuple(archive.infolist())
    except (BadZipFile, OSError) as error:
        message = "下载明细必须是有效 XLSX ZIP 容器"
        raise JsonValueValidationContractError(message) from error


def _validate_xlsx_entry_index(entries: tuple[ZipInfo, ...]) -> None:
    if not 1 <= len(entries) <= _MAX_XLSX_ENTRIES:
        message = "下载明细 ZIP entry 数量超出安全边界"
        raise JsonValueValidationContractError(message)
    names = tuple(entry.filename for entry in entries)
    if len(set(names)) != len(names):
        message = "下载明细 ZIP 不得包含重复 entry"
        raise JsonValueValidationContractError(message)
    required = {"[Content_Types].xml", "xl/workbook.xml"}
    if not required.issubset(names):
        message = "下载明细 ZIP 缺少 XLSX 必需 entry"
        raise JsonValueValidationContractError(message)


def _validated_expanded_total(entry: ZipInfo, current_total: int) -> int:
    name = PurePosixPath(entry.filename.replace("\\", "/"))
    first_part = name.parts[0] if name.parts else ""
    if name.is_absolute() or ".." in name.parts or ":" in first_part or entry.flag_bits & 0x1:
        message = f"下载明细 ZIP 包含不安全 entry: {entry.filename!r}"
        raise JsonValueValidationContractError(message)
    if entry.file_size > _MAX_XLSX_MEMBER_BYTES:
        message = "下载明细 ZIP 单 entry 展开大小超出安全边界"
        raise JsonValueValidationContractError(message)
    expanded_total = current_total + entry.file_size
    if expanded_total > _MAX_XLSX_EXPANDED_BYTES:
        message = "下载明细 ZIP 总展开大小超出安全边界"
        raise JsonValueValidationContractError(message)
    ratio = entry.file_size / max(entry.compress_size, 1)
    if ratio > _MAX_XLSX_COMPRESSION_RATIO:
        message = "下载明细 ZIP 压缩比超出安全边界"
        raise JsonValueValidationContractError(message)
    return expanded_total


def _workbook_text(path: Path) -> str:
    _require_safe_xlsx(path)
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (OSError, ValueError, DefusedXmlException) as error:
        message = "规则库导出文件必须可读取"
        raise JsonValueValidationContractError(message) from error
    try:
        return "\n".join(
            value
            for worksheet in _bounded_worksheets(workbook)
            for value in _worksheet_nonempty_texts(worksheet)
        )
    finally:
        workbook.close()
