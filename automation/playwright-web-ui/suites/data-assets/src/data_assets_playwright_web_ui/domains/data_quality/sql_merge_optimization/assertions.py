"""Pure business assertions for SQL topology and downloaded dirty-data rows."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Final
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from data_assets_playwright_web_ui.domains.data_quality.sql_merge_optimization.model import (
    MergeMode,
    SqlRuleIdentity,
    SqlTopologyExpectation,
    TableSnapshot,
)

if TYPE_CHECKING:
    from pathlib import Path

_INSERT_TARGET_RE: Final = re.compile(
    r"\binsert\s+(?:overwrite\s+table|into(?:\s+table)?)\s+"
    r"((?:`[^`]+`|[A-Za-z_][A-Za-z0-9_$]*)(?:\s*\.\s*(?:`[^`]+`|[A-Za-z_][A-Za-z0-9_$]*))?)",
    flags=re.IGNORECASE,
)
_CREATE_TARGET_RE: Final = re.compile(
    r"\bcreate\s+table\s+(?:if\s+not\s+exists\s+)?"
    r"((?:`[^`]+`|[A-Za-z_][A-Za-z0-9_$]*)(?:\s*\.\s*(?:`[^`]+`|[A-Za-z_][A-Za-z0-9_$]*))?)",
    flags=re.IGNORECASE,
)
_ZERO: Final = Decimal(0)
_MAX_WORKBOOK_BYTES: Final = 10 * 1024 * 1024
_MAX_ARCHIVE_ENTRIES: Final = 512
_MAX_ARCHIVE_ENTRY_BYTES: Final = 20 * 1024 * 1024
_MAX_ARCHIVE_UNCOMPRESSED_BYTES: Final = 40 * 1024 * 1024
_MAX_ARCHIVE_COMPRESSION_RATIO: Final = 200
_SUM_CASE_RE: Final = re.compile(r"\bsum\s*\(\s*case\s+when\b", re.IGNORECASE)
_COUNT_DISTINCT_RE: Final = re.compile(r"\bcount\s*\(\s*distinct\b", re.IGNORECASE)


class SqlMergeAssertionError(AssertionError):
    """Raised when UI-read SQL or downloaded rows violate canonical semantics."""


@dataclass(frozen=True, slots=True)
class SqlTopologyReadback:
    """Sanitized SQL facts read from the task-detail editor."""

    effective_scan_count: int
    source_setup_scan_count: int
    dirty_targets: tuple[str, ...]
    sampling_enabled: bool
    sum_case_count: int
    distinct_count: int
    rule_groups: tuple[tuple[int, ...], ...]
    group_targets: tuple[str, ...]

    def as_json(self) -> dict[str, object]:
        """Return safe business-record values without persisting SQL text."""
        return {
            "effective_scan_count": self.effective_scan_count,
            "source_setup_scan_count": self.source_setup_scan_count,
            "dirty_target_count": len(self.dirty_targets),
            "sampling_enabled": self.sampling_enabled,
            "sum_case_count": self.sum_case_count,
            "distinct_count": self.distinct_count,
            "rule_groups": [list(group) for group in self.rule_groups],
            "group_targets": list(self.group_targets),
        }


@dataclass(frozen=True, slots=True)
class _SqlRuleGroup:
    """One result-producing source scan bound to persisted rules."""

    members: tuple[int, ...]
    sql: str


def assert_shared_dirty_table(sql: str) -> str:
    """Require every rendered INSERT statement to use one normalized dirty table."""
    targets = tuple(_normalize_sql_identifier(match) for match in _INSERT_TARGET_RE.findall(sql))
    if not targets:
        message = "rule SQL must contain an INSERT target for dirty-data storage"
        raise SqlMergeAssertionError(message)
    unique_targets = frozenset(targets)
    if len(unique_targets) != 1:
        message = f"rule SQL must share one dirty-data target; found {len(unique_targets)} targets"
        raise SqlMergeAssertionError(message)
    return targets[0]


def assert_sql_topology(
    sql: str,
    *,
    table_name: str,
    rule_count: int,
    expectation: SqlTopologyExpectation,
    rule_identities: tuple[SqlRuleIdentity, ...],
) -> SqlTopologyReadback:
    """Verify rendered Spark SQL against one explicit canonical grouping model."""
    if not sql.strip():
        message = "rule SQL must contain rendered Spark statements"
        raise SqlMergeAssertionError(message)
    if rule_count < 1:
        message = "rule_count must be positive"
        raise ValueError(message)
    expectation.validate_rule_indices(rule_count)
    _validate_rule_identities(rule_identities, rule_count=rule_count)

    statements = _sql_statements(sql)
    effective_statements, source_setup_scan_count, sampling_enabled = _effective_scan_statements(
        statements,
        table_name=table_name,
        expectation=expectation,
        rule_identities=rule_identities,
    )
    effective_scan_count = len(effective_statements)
    _assert_scan_count(effective_scan_count, expectation)
    groups = _assert_rule_group_topology(
        effective_statements,
        expectation=expectation,
        rule_identities=rule_identities,
    )
    sum_case_count, distinct_count = _assert_merge_operators(groups, expectation)
    rule_groups = tuple(group.members for group in groups)
    group_targets = _assert_dirty_target_topology(
        statements,
        expectation=expectation,
        rule_identities=rule_identities,
    )
    dirty_targets = tuple(dict.fromkeys(group_targets))
    _assert_partition(statements, table_name=table_name, expectation=expectation)

    return SqlTopologyReadback(
        effective_scan_count=effective_scan_count,
        source_setup_scan_count=source_setup_scan_count,
        dirty_targets=dirty_targets,
        sampling_enabled=sampling_enabled,
        sum_case_count=sum_case_count,
        distinct_count=distinct_count,
        rule_groups=rule_groups,
        group_targets=group_targets,
    )


def read_xlsx_snapshot(path: Path, *, maximum_rows: int) -> TableSnapshot:
    """Read a bounded XLSX-compatible download into normalized business values."""
    if isinstance(maximum_rows, bool) or maximum_rows < 1:
        message = "maximum_rows must be a positive integer"
        raise ValueError(message)
    payload = _read_bounded_payload(path)
    _validate_xlsx_archive(payload)
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except (BadZipFile, EOFError, KeyError, OSError, ValueError) as error:
        message = "download must be a readable XLSX workbook"
        raise SqlMergeAssertionError(message) from error

    try:
        sheet = workbook.active
        if sheet is None:
            message = "downloaded detail workbook must contain an active worksheet"
            raise SqlMergeAssertionError(message)
        headers: tuple[str, ...] | None = None
        rows: list[tuple[str, ...]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            row = tuple(_normalize_cell(value) for value in raw_row)
            if not any(row):
                continue
            if headers is None:
                if any(not header for header in row):
                    message = "downloaded detail workbook headers must be non-empty"
                    raise SqlMergeAssertionError(message)
                headers = row
                continue
            if len(rows) >= maximum_rows:
                message = (
                    f"downloaded detail workbook must contain at most {maximum_rows} data rows"
                )
                raise SqlMergeAssertionError(message)
            rows.append(row[: len(headers)])
    finally:
        workbook.close()

    if headers is None:
        message = "downloaded detail workbook must contain a header"
        raise SqlMergeAssertionError(message)
    return TableSnapshot(headers=headers, rows=tuple(rows))


def assert_download_matches_visible_rows(
    downloaded: TableSnapshot,
    visible: TableSnapshot,
) -> None:
    """Require downloaded dirty-data content to equal the rendered rows exactly."""
    if downloaded != visible:
        message = "downloaded dirty-data rows must exactly match the visible business table"
        raise SqlMergeAssertionError(message)


def _normalize_sql_identifier(value: str) -> str:
    segments = tuple(segment.strip().strip("`") for segment in value.split("."))
    return ".".join(segments).lower()


def _count_from_table(sql: str, table_name: str) -> int:
    optional_schema = r"(?:(?:`?[A-Za-z_][A-Za-z0-9_$]*`?)\s*\.\s*)?"
    table = rf"`?{re.escape(table_name.lower())}`?"
    return len(re.findall(rf"\bfrom\s+{optional_schema}{table}\b", sql))


def _sql_statements(sql: str) -> tuple[str, ...]:
    statements = tuple(
        " ".join(branch.split())
        for value in sql.split(";")
        for branch in re.split(r"\bunion(?:\s+all)?\b", value, flags=re.IGNORECASE)
        if branch.strip()
    )
    if not statements:
        message = "rule SQL must contain executable statements"
        raise SqlMergeAssertionError(message)
    return statements


def _effective_scan_statements(
    statements: tuple[str, ...],
    *,
    table_name: str,
    expectation: SqlTopologyExpectation,
    rule_identities: tuple[SqlRuleIdentity, ...],
) -> tuple[tuple[str, ...], int, bool]:
    sample_table = f"{table_name}_temp_sample_table"
    lowered = tuple(statement.lower() for statement in statements)
    result_statements = tuple(
        statement
        for statement in lowered
        if _structured_result_rule_members(statement, rule_identities)
    )
    source_statements = tuple(
        statement for statement in result_statements if _count_from_table(statement, table_name) > 0
    )
    sample_statements = tuple(
        statement
        for statement in result_statements
        if _count_from_table(statement, sample_table) > 0
    )
    sample_setup_statements = tuple(
        statement
        for statement in lowered
        if _is_target(statement, sample_table) and _count_from_table(statement, table_name) > 0
    )
    source_setup_count = sum(
        _count_from_table(statement, table_name) for statement in sample_setup_statements
    )
    sample_scan_count = sum(
        _count_from_table(statement, sample_table) for statement in sample_statements
    )
    has_rand = any(
        re.search(r"\brand\s*\(", statement) is not None for statement in sample_setup_statements
    )
    sampling_enabled = bool(sample_setup_statements) and has_rand
    if expectation.sampling_percent is not None:
        ratio_pattern = re.compile(r"\brand\s*\(\s*\)\s*(?:<=|<)\s*(?:0?\.5(?:0+)?|50\s*/\s*100)\b")
        if (
            not sampling_enabled
            or source_setup_count != 1
            or sample_scan_count < 1
            or not any(ratio_pattern.search(statement) for statement in sample_setup_statements)
        ):
            message = (
                "sampling topology must use an exact 50% rand() predicate, "
                "one source setup, and the case temp table"
            )
            raise SqlMergeAssertionError(message)
        if any(_count_from_table(statement, sample_table) != 1 for statement in sample_statements):
            message = "each effective sampled rule group must scan the temp table exactly once"
            raise SqlMergeAssertionError(message)
        return sample_statements, source_setup_count, sampling_enabled
    if any(re.search(r"\brand\s*\(", statement) for statement in lowered) or any(
        sample_table.lower() in statement for statement in lowered
    ):
        message = "non-sampling topology must not use rand() or a temp sample table"
        raise SqlMergeAssertionError(message)
    if any(_count_from_table(statement, table_name) != 1 for statement in source_statements):
        message = "each effective rule group must scan the source table exactly once"
        raise SqlMergeAssertionError(message)
    source_scan_count = sum(
        _count_from_table(statement, table_name) for statement in source_statements
    )
    return source_statements, source_scan_count, sampling_enabled


def _assert_scan_count(
    effective_scan_count: int,
    expectation: SqlTopologyExpectation,
) -> None:
    expected = expectation.expected_scan_groups
    if expectation.mode is MergeMode.FULL and effective_scan_count != expected:
        message = "fully merged topology must scan its effective source exactly once"
        raise SqlMergeAssertionError(message)
    if expectation.mode is not MergeMode.FULL and effective_scan_count < expected:
        label = "independent" if expectation.mode is MergeMode.NONE else "partial"
        message = f"{label} topology must preserve every declared scan group"
        raise SqlMergeAssertionError(message)


def _assert_merge_operators(
    groups: tuple[_SqlRuleGroup, ...],
    expectation: SqlTopologyExpectation,
) -> tuple[int, int]:
    merged_groups = {tuple(sorted(group)) for group in expectation.merged_rule_groups}
    distinct_groups = {tuple(sorted(group)) for group in expectation.distinct_rule_groups}
    sum_case_count = 0
    distinct_count = 0
    for group in groups:
        group_sum_count = len(_SUM_CASE_RE.findall(group.sql))
        group_distinct_count = len(_COUNT_DISTINCT_RE.findall(group.sql))
        sum_case_count += group_sum_count
        distinct_count += group_distinct_count
        if group.members in merged_groups and group_sum_count < 1:
            message = (
                "every multi-rule SQL group must contain its own SUM(CASE WHEN) merge signature"
            )
            raise SqlMergeAssertionError(message)
        if group.members in distinct_groups and group_distinct_count < 1:
            message = "declared SQL rule group must retain count(DISTINCT) semantics"
            raise SqlMergeAssertionError(message)
    return sum_case_count, distinct_count


def _dirty_targets(sql: str) -> tuple[str, ...]:
    targets = tuple(
        normalized
        for match in (*_INSERT_TARGET_RE.findall(sql), *_CREATE_TARGET_RE.findall(sql))
        if not (normalized := _normalize_sql_identifier(match)).endswith("_temp_sample_table")
    )
    return tuple(dict.fromkeys(targets))


def _validate_rule_identities(
    identities: tuple[SqlRuleIdentity, ...],
    *,
    rule_count: int,
) -> None:
    indices = tuple(identity.index for identity in identities)
    tokens = tuple(identity.token.lower() for identity in identities)
    if set(indices) != set(range(1, rule_count + 1)) or len(indices) != rule_count:
        message = "persisted rule identities must cover every configured rule exactly once"
        raise SqlMergeAssertionError(message)
    if len(set(tokens)) != rule_count:
        message = "persisted rule identity tokens must be unique"
        raise SqlMergeAssertionError(message)


def _assert_rule_group_topology(
    effective_statements: tuple[str, ...],
    *,
    expectation: SqlTopologyExpectation,
    rule_identities: tuple[SqlRuleIdentity, ...],
) -> tuple[_SqlRuleGroup, ...]:
    expected_groups = expectation.expected_rule_groups
    actual_groups: list[_SqlRuleGroup] = []
    seen_indices: set[int] = set()
    for statement in effective_statements:
        lowered = statement.lower()
        members = _structured_result_rule_members(lowered, rule_identities)
        if not members:
            message = (
                "every effective source scan must bind rule IDs at a supported generated-SQL "
                "structure"
            )
            raise SqlMergeAssertionError(message)
        if seen_indices & set(members):
            message = "a persisted rule identity cannot belong to multiple SQL scan groups"
            raise SqlMergeAssertionError(message)
        seen_indices.update(members)
        actual_groups.append(_SqlRuleGroup(members=members, sql=lowered))

    if sorted(group.members for group in actual_groups) != sorted(expected_groups):
        message = "rendered SQL rule-group members must exactly match canonical topology"
        raise SqlMergeAssertionError(message)
    return tuple(actual_groups)


def _structured_result_rule_members(
    sql: str,
    identities: tuple[SqlRuleIdentity, ...],
) -> tuple[int, ...]:
    """Match rule IDs only at source-backed generated-SQL identity positions."""
    members: list[int] = []
    rule_id_cases = tuple(
        match.group("body")
        for match in re.finditer(
            r"\bcase\b(?P<body>.*?)\bend\s+as\s+rule_id\b",
            sql,
            flags=re.IGNORECASE,
        )
    )
    for identity in identities:
        token = re.escape(identity.token)
        result_patterns = (
            rf"\bhit_cnt_rule_{token}\b",
            rf"\bhit_rule_expansion_{token}\b",
            rf"(?<![A-Za-z0-9_]){token}\s+as\s+rule_id\b",
        )
        in_rule_id_case = any(
            re.search(rf"\bthen\s+{token}(?![0-9])", body, re.IGNORECASE) is not None
            for body in rule_id_cases
        )
        if in_rule_id_case or any(
            re.search(pattern, sql, re.IGNORECASE) is not None for pattern in result_patterns
        ):
            members.append(identity.index)
    return tuple(sorted(members))


def _assert_dirty_target_topology(
    statements: tuple[str, ...],
    *,
    expectation: SqlTopologyExpectation,
    rule_identities: tuple[SqlRuleIdentity, ...],
) -> tuple[str, ...]:
    """Bind only Spark dirty CREATE/INSERT operations to exact rule groups."""
    target_members: dict[str, set[int]] = {}
    for statement in statements:
        targets = _dirty_targets(statement)
        members = _structured_dirty_rule_members(statement, targets, rule_identities)
        if not members:
            continue
        if len(targets) != 1:
            message = "one dirty-data operation must bind exactly one target"
            raise SqlMergeAssertionError(message)
        target_members.setdefault(targets[0], set()).update(members)

    actual_by_members = {
        tuple(sorted(members)): target for target, members in target_members.items()
    }
    if len(actual_by_members) != len(target_members):
        message = "different dirty-data targets cannot represent the same rule group"
        raise SqlMergeAssertionError(message)
    expected_target_groups = tuple(expectation.expected_dirty_target_groups)
    if sorted(actual_by_members) != sorted(expected_target_groups):
        message = "dirty-data target equivalence must exactly match canonical SQL rule groups"
        raise SqlMergeAssertionError(message)
    return tuple(
        next(target for members, target in actual_by_members.items() if set(group) <= set(members))
        for group in expectation.expected_rule_groups
    )


def _structured_dirty_rule_members(
    sql: str,
    targets: tuple[str, ...],
    identities: tuple[SqlRuleIdentity, ...],
) -> tuple[int, ...]:
    """Recognize merged explode tags or the isolated target suffix contract."""
    members: list[int] = []
    for identity in identities:
        token = re.escape(identity.token)
        merged_marker = re.search(
            rf"['\"]{token}['\"]\s*,\s*null\s*\)",
            sql,
            re.IGNORECASE,
        )
        isolated_target = any(
            re.search(rf"_{token}$", target, re.IGNORECASE) is not None for target in targets
        )
        if merged_marker is not None or isolated_target:
            members.append(identity.index)
    return tuple(sorted(members))


def _is_target(sql: str, table_name: str) -> bool:
    expected = table_name.lower()
    targets = (*_INSERT_TARGET_RE.findall(sql), *_CREATE_TARGET_RE.findall(sql))
    return any(_normalize_sql_identifier(target).endswith(expected) for target in targets)


def _assert_partition(
    statements: tuple[str, ...],
    *,
    table_name: str,
    expectation: SqlTopologyExpectation,
) -> None:
    partition_pattern = re.compile(r"\bdt\s*=\s*['\"]2026-08-04['\"]")
    sample_table = f"{table_name}_temp_sample_table"
    lowered = tuple(statement.lower() for statement in statements)
    source_scans = tuple(
        statement for statement in lowered if _count_from_table(statement, table_name) > 0
    )
    effective_scans = (
        tuple(statement for statement in lowered if _count_from_table(statement, sample_table) > 0)
        if expectation.sampling_percent is not None
        else source_scans
    )
    partition_bound_scans = (
        source_scans if expectation.sampling_percent is not None else effective_scans
    )
    if expectation.partition_filter is not None and (
        not partition_bound_scans
        or any(partition_pattern.search(statement) is None for statement in partition_bound_scans)
    ):
        message = "every canonical source scan must retain dt=2026-08-04"
        raise SqlMergeAssertionError(message)
    if expectation.partition_filter is None and any(
        partition_pattern.search(statement) is not None
        for statement in (*source_scans, *effective_scans)
    ):
        message = "unpartitioned topology must not inject the canonical dt filter"
        raise SqlMergeAssertionError(message)


def _read_bounded_payload(path: Path) -> bytes:
    try:
        expected_size = path.stat().st_size
        if expected_size <= 0 or expected_size > _MAX_WORKBOOK_BYTES:
            message = "downloaded detail workbook violates the fixed size limit"
            raise SqlMergeAssertionError(message)
        with path.open("rb") as stream:
            payload = stream.read(_MAX_WORKBOOK_BYTES + 1)
    except OSError as error:
        message = "download must be a readable XLSX workbook"
        raise SqlMergeAssertionError(message) from error
    if len(payload) != expected_size or len(payload) > _MAX_WORKBOOK_BYTES:
        message = "downloaded detail workbook changed during bounded reading"
        raise SqlMergeAssertionError(message)
    return payload


def _validate_xlsx_archive(payload: bytes) -> None:
    """Reject oversized or suspiciously compressed XLSX members before XML parsing."""
    try:
        with ZipFile(BytesIO(payload)) as archive:
            entries = tuple(entry for entry in archive.infolist() if not entry.is_dir())
    except (BadZipFile, EOFError, OSError, ValueError) as error:
        message = "download must be a readable XLSX workbook"
        raise SqlMergeAssertionError(message) from error
    total_uncompressed = 0
    if not entries or len(entries) > _MAX_ARCHIVE_ENTRIES:
        message = "downloaded detail workbook violates archive safety limits"
        raise SqlMergeAssertionError(message)
    for entry in entries:
        total_uncompressed += entry.file_size
        ratio = entry.file_size / max(entry.compress_size, 1)
        if (
            entry.flag_bits & 0x1
            or entry.file_size > _MAX_ARCHIVE_ENTRY_BYTES
            or total_uncompressed > _MAX_ARCHIVE_UNCOMPRESSED_BYTES
            or ratio > _MAX_ARCHIVE_COMPRESSION_RATIO
        ):
            message = "downloaded detail workbook violates archive safety limits"
            raise SqlMergeAssertionError(message)


def _normalize_cell(value: object) -> str:
    if value is None:
        result = ""
    elif isinstance(value, bool):
        result = "true" if value else "false"
    elif isinstance(value, datetime):
        result = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        result = value.isoformat()
    elif isinstance(value, float):
        decimal = Decimal(str(value))
        result = _decimal_text(decimal)
    elif isinstance(value, Decimal):
        result = _decimal_text(value)
    else:
        result = str(value).strip()
    return result


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == _ZERO:
        return "0"
    return format(normalized, "f")
