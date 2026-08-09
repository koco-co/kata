"""Safe XLSX builders and readers for JSON configuration import/export journeys."""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING, Final, Protocol, cast
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook

if TYPE_CHECKING:
    from pathlib import Path

    from data_assets_playwright_web_ui.domains.data_quality.json_configuration.model import (
        JsonImportRow,
    )


class _ColorLike(Protocol):
    type: str
    rgb: str | None


class _FillLike(Protocol):
    fgColor: _ColorLike  # noqa: N815


class _CommentLike(Protocol):
    text: str


class _InspectableCell(Protocol):
    value: object
    coordinate: str
    fill: _FillLike
    comment: _CommentLike | None


IMPORT_SHEET_NAMES: Final = ("一层", "二层", "三层", "四层", "五层")
_PARENT_HEADERS: Final = (
    "第一层级key名",
    "第二层级key名",
    "第三层级key名",
    "第四层级key名",
)
_BASE_HEADERS: Final = ("* key", "中文名称", "value格式")
_RED_RGB_VALUES: Final = frozenset({"FF0000", "FFFF0000", "00FF0000"})
_MAX_XLSX_FILE_BYTES: Final = 20 * 1024 * 1024
_MAX_XLSX_EXPANDED_BYTES: Final = 100 * 1024 * 1024
_MAX_XLSX_MEMBER_BYTES: Final = 50 * 1024 * 1024
_MAX_XLSX_MEMBERS: Final = 1_024


@dataclass(frozen=True, slots=True)
class TemplateInspection:
    """Sheet names and header order read from a downloaded template."""

    sheet_names: tuple[str, ...]
    headers: tuple[tuple[str, ...], ...]


@dataclass(frozen=True, slots=True)
class ErrorCellInspection:
    """Cell location, red-fill state, and comment from an import error workbook."""

    coordinate: str
    is_red: bool
    comment: str | None


@dataclass(frozen=True, slots=True)
class ExportInspection:
    """Ordered headers and named business rows from a UI export."""

    headers: tuple[str, ...]
    rows: tuple[dict[str, str], ...]


class WorkbookContractError(AssertionError):
    """Raised when a downloaded workbook cannot satisfy the canonical contract."""


class JsonConfigurationWorkbook:
    """Build input workbooks and inspect files downloaded through the UI."""

    @staticmethod
    def headers_for_level(level: int) -> tuple[str, ...]:
        """Return the exact import template header sequence for one level."""
        if isinstance(level, bool) or not 1 <= level <= len(IMPORT_SHEET_NAMES):
            message = "level must be an integer from 1 through 5"
            raise ValueError(message)
        parents = tuple(f"* {header}" for header in _PARENT_HEADERS[: level - 1])
        return (*parents, *_BASE_HEADERS)

    @classmethod
    def build_import(cls, path: Path, *, rows: tuple[JsonImportRow, ...]) -> Path:
        """Create one deterministic five-sheet XLSX fixture under pytest's temp directory."""
        if path.suffix.lower() != ".xlsx":
            message = "import workbook path must use the .xlsx suffix"
            raise ValueError(message)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        active = workbook.active
        if active is None:
            message = "new workbook must expose its initial worksheet"
            raise RuntimeError(message)
        workbook.remove(active)
        for level, name in enumerate(IMPORT_SHEET_NAMES, start=1):
            sheet = workbook.create_sheet(name)
            sheet.append(cls.headers_for_level(level))
            for row in rows:
                if row.level == level:
                    sheet.append(row.values)
        workbook.save(path)
        return path

    @classmethod
    def inspect_template(cls, path: Path) -> TemplateInspection:
        """Read a downloaded template and return its ordered five-level schema."""
        _require_safe_workbook(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            names = tuple(workbook.sheetnames)
            headers = tuple(
                tuple(_cell_text(cell.value) for cell in workbook[name][1]) for name in names
            )
            return TemplateInspection(sheet_names=names, headers=headers)
        finally:
            workbook.close()

    @staticmethod
    def inspect_error_cell(
        path: Path,
        *,
        sheet_name: str,
        value: str,
    ) -> ErrorCellInspection:
        """Find an exact value and report the product's red annotation on that cell."""
        _require_safe_workbook(path)
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                message = f'error workbook must contain sheet "{sheet_name}"'
                raise WorkbookContractError(message)
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for raw_cell in row:
                    cell = cast("_InspectableCell", raw_cell)
                    if _cell_text(cell.value) != value:
                        continue
                    color = cell.fill.fgColor
                    rgb = color.rgb if color.type == "rgb" else None
                    comment = None if cell.comment is None else cell.comment.text.strip()
                    return ErrorCellInspection(
                        coordinate=cell.coordinate,
                        is_red=isinstance(rgb, str) and rgb.upper() in _RED_RGB_VALUES,
                        comment=comment,
                    )
            message = f'error workbook sheet "{sheet_name}" must contain exact value "{value}"'
            raise WorkbookContractError(message)
        finally:
            workbook.close()

    @staticmethod
    def inspect_error_coordinate(
        path: Path,
        *,
        sheet_name: str,
        coordinate: str,
    ) -> ErrorCellInspection:
        """Inspect one canonical error coordinate, including blank required cells."""
        _require_safe_workbook(path)
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                message = f'error workbook must contain sheet "{sheet_name}"'
                raise WorkbookContractError(message)
            raw_cell = workbook[sheet_name][coordinate]
            cell = cast("_InspectableCell", raw_cell)
            color = cell.fill.fgColor
            rgb = color.rgb if color.type == "rgb" else None
            comment = None if cell.comment is None else cell.comment.text.strip()
            return ErrorCellInspection(
                coordinate=cell.coordinate,
                is_red=isinstance(rgb, str) and rgb.upper() in _RED_RGB_VALUES,
                comment=comment,
            )
        finally:
            workbook.close()

    @staticmethod
    def inspect_export(path: Path) -> ExportInspection:
        """Read ordered headers and named rows from the first exported sheet."""
        _require_safe_workbook(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                message = "export workbook must contain a header row"
                raise WorkbookContractError(message)
            headers = tuple(_cell_text(value) for value in raw_rows[0])
            if not headers or any(not header for header in headers):
                message = "export workbook headers must be non-empty"
                raise WorkbookContractError(message)
            rows: list[dict[str, str]] = []
            for raw_row in raw_rows[1:]:
                values = tuple(_cell_text(value) for value in raw_row[: len(headers)])
                if not any(values):
                    continue
                rows.append(dict(zip(headers, values, strict=True)))
            return ExportInspection(headers=headers, rows=tuple(rows))
        finally:
            workbook.close()

    @staticmethod
    def read_export(path: Path) -> tuple[dict[str, str], ...]:
        """Read the first exported sheet as named non-blank business rows."""
        return JsonConfigurationWorkbook.inspect_export(path).rows


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _require_safe_workbook(path: Path) -> None:
    """Reject oversized or suspicious XLSX archives before openpyxl parses XML."""
    try:
        file_size = path.stat().st_size
    except OSError as error:
        message = "workbook must be a readable local file"
        raise WorkbookContractError(message) from error
    if not 0 < file_size <= _MAX_XLSX_FILE_BYTES:
        message = "workbook exceeds the safe file-size limit"
        raise WorkbookContractError(message)
    try:
        with ZipFile(path) as archive:
            members = archive.infolist()
    except (BadZipFile, OSError) as error:
        message = "workbook must be a valid XLSX archive"
        raise WorkbookContractError(message) from error
    if len(members) > _MAX_XLSX_MEMBERS:
        message = "workbook archive contains too many entries"
        raise WorkbookContractError(message)
    expanded_sizes = tuple(member.file_size for member in members)
    if (
        any(size > _MAX_XLSX_MEMBER_BYTES for size in expanded_sizes)
        or sum(expanded_sizes) > _MAX_XLSX_EXPANDED_BYTES
    ):
        message = "workbook archive exceeds the safe expanded-size limit"
        raise WorkbookContractError(message)
